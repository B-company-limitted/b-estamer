# Copyright (c) 2026 BRUNO CONSTRUCTION EMPIRE LTD
# B-ESTAMER V5.7 - 20 ELEMENTS + EDITABLE RATES + PDF UPLOAD | 0787993679

import streamlit as st
import pandas as pd
import plotly.express as px
from io import BytesIO
import math
from PIL import Image
import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, Alignment, PatternFill

st.set_page_config(page_title="B-ESTAMER 5.7", layout="wide")  
st.title("B-ESTAMER 5.7 ULTIMATE 🏗️")
st.caption("20 Elements + Editable Rates + PDF Upload + Auto Excel")

import streamlit as st

# 1. Bika selection muri session_state
if 'material' not in st.session_state:
    st.session_state.material = 'Steel'
if 'subtype' not in st.session_state:
    st.session_state.subtype = ''

# 2. Material ya mbere
materials = ['Steel', 'Concrete', 'Timber', 'Aluminum']
st.session_state.material = st.selectbox(
    'Hitamo Material',
    materials,
    index=materials.index(st.session_state.material),
    key='material_select'
)

# 3. Subtypes zihinduka hashingiye kuri material
subtypes_dict = {
    'Steel': ['Mild Steel', 'High Carbon Steel', 'Stainless Steel'],
    'Concrete': ['C25/30', 'C30/37', 'C40/50'],
    'Timber': ['Softwood', 'Hardwood', 'Plywood'],
    'Aluminum': ['6061-T6', '7075-T6']
} # Material ya mbere
st.session_state.material = st.selectbox(
    "Hitamo Material", 
    list(subtypes_dict.keys()),
    key="material_select"
)

# Subtype ya kabiri - IBIKOMEYE HANO
subtypes = subtypes_dict[st.session_state.material]

# Reba niba subtype ya kera ikiri muri list nshya
if st.session_state.subtype not in subtypes:
    st.session_state.subtype = subtypes[0]

# Ubu ukoreshe index + key zitandukanye
subtype_index = subtypes.index(st.session_state.subtype)
st.session_state.subtype = st.selectbox(
    "Hitamo Subtype",
    subtypes,
    index=subtype_index, # <- Iki gituma idasubira kuri 0
    key="subtype_select" # <- Key itandukanye na material
)
# Material ya mbere
st.session_state.material = st.selectbox(
    "Hitamo Material", 
    list(subtypes_dict.keys()),
    key="material_select"
)

# Subtype ya kabiri - IBIKOMEYE HANO
subtypes = subtypes_dict[st.session_state.material]

# Reba niba subtype ya kera ikiri muri list nshya
if st.session_state.subtype not in subtypes:
    st.session_state.subtype = subtypes[0]
subtype_index = subtypes.index(st.session_state.subtype) if st.session_state.subtype in subtypes else 0

st.session_state.subtype = st.selectbox(
    "Hitamo Subtype",
    subtypes,
    index=subtype_index,
    key="subtype_select"
)
# Ubu ukoreshe index + key zitandukanye
subtype_index = subtypes.index(st.session_state.subtype)

st.write(f"Uhitamo: {st.session_state.material} → {st.session_state.subtype}")
### Amakosa 3 abantu bakora:
Zombi selectbox zifite key imwe	Buri selectbox igomba kugira key yihariye
index=0 buri gihe	Koresha index=materials.index(st.session_state.material)
Nta session_state	Bika selection muri st.session_state
### Trick yihuse:
Niba ushaka ko subtype ihinduka iyo material ihindutse:
if 'last_material' not in st.session_state:
    st.session_state.last_material = st.session_state.material

if st.session_state.material!= st.session_state.last_material:
    st.session_state.subtype = subtypes[0] # Subira ku wa mbere
    st.session_state.last_material = st.session_state.material
# ============= MATERIAL TREE - MAIN + SUB TYPES =============
MATERIAL_TREE = {
    "CEMENT": ["CEM II 42.5N","CEM I 52.5N","White Cement","Quick Set Cement","Portland Cement"],
    "SAND": ["River Sand","Crushed Sand","Plaster Sand","Filling Sand","Quarry Dust"],
    "STONE": ["Stone 20mm","Stone 40mm","Ballast","Hardcore","Aggregate Dust","Chippings"],
    "BRICK": ["Clay Brick 6 inch","Clay Brick 4 inch","Concrete Block 23cm Solid","Concrete Block 15cm Solid","Concrete Block 23cm Hollow","Concrete Block 15cm Hollow","Paving Block","Face Brick"],
    "STEEL_BAR": ["R6 Stirrups","R8 Stirrups","R10 Steel","Y12 Steel","Y16 Steel","Y20 Steel","Y25 Steel","Y32 Steel","Y40 Steel","Y50 Steel"],
    "TIMBER": ["Cypress 2x2","Cypress 3x2","Cypress 4x2","Plywood 18mm","Marine Board 18mm","MDF Board","Hardwood","Softwood","Pine"],
    "IRON_SHEET": ["Mabati G30 Versatile","Mabati G28 IT4","Mabati G30 Corrugated","Resincot","Gauge 28","Gauge 30","Aluminium Sheet"],
    "PAINT": ["Emulsion Paint","Weather Guard","Undercoat","Varnish","Oil Based","Water Based","Primer"],
    "TILES": ["Ceramic Tile 30x30","Ceramic Tile 60x60","Porcelain Tile","Granite Tile","Marble Tile","PVC Tile","Terrazzo"],
    "CONCRETE": ["Concrete C15","Concrete C20","Concrete C25","Concrete C30","Concrete C35","Concrete C40","Ready Mix C25","Ready Mix C30"],
    "GLASS": ["Glass 4mm Clear","Glass 5mm Tinted","Glass 6mm Laminated","Mirror","Frosted Glass","Bullet Proof"],
    "PIPES": ["PPR 20mm","PPR 25mm","PPR 32mm","PPR 40mm","PVC 4inch Waste","PVC 2inch Waste","PVC 6inch","HDPE 50mm","GI Pipe 1inch"],
    "WIRES": ["Cable 1.5mm","Cable 2.5mm","Cable 4mm","Cable 6mm","Cable 10mm","Cable 16mm","Twin Earth","Armoured Cable"],
    "DOORS": ["Steel Door 90x210","Flush Door","Panel Door","PVC Door","Glass Door","Security Door"],
    "WINDOWS": ["Aluminium Sliding","Aluminium Casement","Louver Window","Wooden Window","UPVC Window"],
    "WATERPROOF": ["APP Membrane","Bitumen Paint","Cementitious Waterproof","Liquid Membrane","EPDM"],
    "CEILING": ["Gypsum Board","PVC Ceiling","Acoustic Ceiling","Wooden T&G","Suspended Ceiling","Mineral Fiber"],
    "NAILS": ["Nails 1 inch","Nails 2 inch","Nails 3 inch","Nails 4 inch","Roofing Nails","Concrete Nails","U-Nails"],
    "BINDING": ["Binding Wire","BRC Mesh A142","BRC Mesh A98","Chain Link","Barbed Wire"],
    "FITTINGS": ["WC Complete","Wash Basin","Kitchen Sink","Shower Mixer","Towel Rail"],
    "ELECTRICAL": ["Socket 13A","Light Point","Switch","DB Board","Earth Rod"],
}

# ============= 20 ELEMENTS DATABASE =============
ELEMENTS = {
    "a. Site Works": {
        "items": ["Excavation", "Hardcore", "Concrete Bed"],
        "inputs": ["Length m", "Width m", "Depth m"],
        "formula": lambda l,w,d: {
            "Excavation": {"qty": l*w*d, "unit":"m3", "rate":3500},
            "Hardcore": {"qty": l*w*0.3, "unit":"m3", "rate":20000},
            "Concrete Bed C15": {"qty": l*w*0.05, "unit":"m3", "rate":160000}
        }
    },
    "b. Concrete Footing": {
        "items": ["Concrete C25", "Steel Y12", "Formwork"],
        "inputs": ["Length m", "Width m", "Depth m", "Number"],
        "formula": lambda l,w,d,n: {
            "Concrete C25": {"qty": l*w*d*n, "unit":"m3", "rate":180000},
            "Steel Y12": {"qty": l*w*d*n*80, "unit":"kg", "rate":1300},
            "Formwork": {"qty": 2*(l+w)*d*n, "unit":"m2", "rate":8500}
        }
    },
    "c. Blockwork": {
        "items": ["Blocks 6 inch", "Mortar 1:4"],
        "inputs": ["Wall Length m", "Wall Height m", "Block Type"],
        "formula": lambda l,h,typ: {
            "Clay Brick 6 inch" if typ=="6 inch" else "Clay Brick 4 inch": {"qty": l*h*(60 if typ=="6 inch" else 40)*1.05, "unit":"pc", "rate":250 if typ=="6 inch" else 180},
            "Mortar 1:4": {"qty": l*h*0.025, "unit":"m3", "rate":95000}
        }
    },
    "d. Floors": {
        "items": ["Concrete Slab C25", "BRC Mesh A142", "Screed 1:4"],
        "inputs": ["Length m", "Width m", "Thickness m"],
        "formula": lambda l,w,t: {
            "Concrete C25": {"qty": l*w*t, "unit":"m3", "rate":180000},
            "BRC Mesh A142": {"qty": l*w*1.05, "unit":"m2", "rate":1700},
            "Screed 1:4": {"qty": l*w, "unit":"m2", "rate":8500}
        }
    },
    "e. Roof": {
        "items": ["Mabati G30", "Timber 3x2", "Roofing Nails"],
        "inputs": ["Roof Area m2", "Slope Factor"],
        "formula": lambda a,sf: {
            "Mabati G30 Versatile": {"qty": a*sf*1.1, "unit":"m2", "rate":8500},
            "Cypress 3x2": {"qty": a*sf*2.5, "unit":"m", "rate":650},
            "Roofing Nails": {"qty": a*sf*0.3, "unit":"kg", "rate":2500}
        }
    },
    "f. Finishing": {
        "items": ["Plaster 1:4", "Ceiling Gypsum", "Emulsion Paint"],
        "inputs": ["Wall Area m2", "Ceiling Area m2"],
        "formula": lambda w,c: {
            "Plaster 1:4": {"qty": w+c, "unit":"m2", "rate":4500},
            "Gypsum Board": {"qty": c/2.88, "unit":"pc", "rate":18000},
            "Emulsion Paint": {"qty": (w+c)*0.15, "unit":"L", "rate":8500}
        }
    },
    "g. Doors": {
        "items": ["Steel Door", "Flush Door"],
        "inputs": ["Steel Doors pcs", "Flush Doors pcs"],
        "formula": lambda s,f: {
            "Steel Door 90x210": {"qty": s, "unit":"pc", "rate":95000},
            "Flush Door": {"qty": f, "unit":"pc", "rate":45000}
        }
    },
    "h. Windows": {
        "items": ["Aluminium Window"],
        "inputs": ["Total Window Area m2"],
        "formula": lambda a: {
            "Aluminium Sliding": {"qty": a, "unit":"m2", "rate":65000}
        }
    },
    "i. Fittings": {
        "items": ["WC Complete", "Wash Basin", "Kitchen Sink"],
        "inputs": ["WC pcs", "Basin pcs", "Sink pcs"],
        "formula": lambda wc,b,s: {
            "WC Complete": {"qty": wc, "unit":"pc", "rate":180000},
            "Wash Basin": {"qty": b, "unit":"pc", "rate":95000},
            "Kitchen Sink": {"qty": s, "unit":"pc", "rate":75000}
        }
    },
    "j. Plumbing Installation": {
        "items": ["PPR 20mm", "PPR 25mm", "PPR Fittings"],
        "inputs": ["PPR 20mm m", "PPR 25mm m"],
        "formula": lambda p20,p25: {
            "PPR 20mm": {"qty": p20, "unit":"m", "rate":2500},
            "PPR 25mm": {"qty": p25, "unit":"m", "rate":3500},
            "PPR Fittings": {"qty": (p20+p25)*0.3, "unit":"pc", "rate":1500}
        }
    },
    "k. Drainage Work": {
        "items": ["PVC 4inch", "Manholes"],
        "inputs": ["PVC Length m", "Manholes pcs"],
        "formula": lambda l,m: {
            "PVC 4inch Waste": {"qty": l, "unit":"m", "rate":4500},
            "Manhole": {"qty": m, "unit":"pc", "rate":125000}
        }
    },
    "l. External Works": {
        "items": ["Road Tarmac", "Chain Link Fence", "Grass"],
        "inputs": ["Road Area m2", "Fence Length m", "Grass Area m2"],
        "formula": lambda r,f,g: {
            "Road Tarmac": {"qty": r, "unit":"m2", "rate":35000},
            "Chain Link Fence": {"qty": f, "unit":"m", "rate":8500},
            "Grass": {"qty": g, "unit":"m2", "rate":2500}
        }
    },
    "m. Columns": {
        "items": ["Concrete C30", "Steel Y16", "Formwork", "Stirrups R8"],
        "inputs": ["Length m", "Width m", "Height m", "Number", "Main Bar Dia"],
        "formula": lambda l,w,h,n,dia: {
            "Concrete C30": {"qty": l*w*h*n, "unit":"m3", "rate":210000},
            "Y16 Steel" if dia==16 else "Y20 Steel": {"qty": l*w*h*n*120, "unit":"kg", "rate":1250 if dia==16 else 1260},
            "Formwork": {"qty": 2*(l+w)*h*n, "unit":"m2", "rate":8500},
            "R8 Stirrups": {"qty": l*w*h*n*15, "unit":"kg", "rate":1320}
        }
    },
    "n. Beams": {
        "items": ["Concrete C30", "Steel Y16", "Formwork", "Stirrups R8"],
        "inputs": ["Length m", "Width m", "Depth m", "Number", "Main Bar Dia"],
        "formula": lambda l,w,d,n,dia: {
            "Concrete C30": {"qty": l*w*d*n, "unit":"m3", "rate":210000},
            "Y16 Steel" if dia==16 else "Y20 Steel": {"qty": l*w*d*n*100, "unit":"kg", "rate":1250 if dia==16 else 1260},
            "Formwork": {"qty": (2*d+w)*l*n, "unit":"m2", "rate":8500},
            "R8 Stirrups": {"qty": l*w*d*n*12, "unit":"kg", "rate":1320}
        }
    },
    "o. Stairs": {
        "items": ["Concrete C25", "Steel Y12", "Formwork", "Terrazzo"],
        "inputs": ["Width m", "Total Rise m", "Going m", "Steps Number"],
        "formula": lambda w,tr,g,n: {
            "Concrete C25": {"qty": w*tr*g*n*0.5, "unit":"m3", "rate":180000},
            "Y12 Steel": {"qty": w*tr*g*n*0.5*80, "unit":"kg", "rate":1300},
            "Formwork": {"qty": w*tr*n*1.5, "unit":"m2", "rate":8500},
            "Terrazzo": {"qty": w*g*n, "unit":"m2", "rate":35000}
        }
    },
    "p. Lintels": {
        "items": ["Concrete C20", "Steel Y12", "Formwork"],
        "inputs": ["Length m", "Width m", "Depth m", "Number"],
        "formula": lambda l,w,d,n: {
            "Concrete C20": {"qty": l*w*d*n, "unit":"m3", "rate":170000},
            "Y12 Steel": {"qty": l*w*d*n*60, "unit":"kg", "rate":1300},
            "Formwork": {"qty": (2*d+w)*l*n, "unit":"m2", "rate":8500}
        }
    },
    "q. Septic Tank": {
        "items": ["Excavation", "Concrete C25", "Blocks 6 inch", "Cover Slab"],
        "inputs": ["Length m", "Width m", "Depth m"],
        "formula": lambda l,w,d: {
            "Excavation": {"qty": l*w*d*1.2, "unit":"m3", "rate":3500},
            "Concrete C25": {"qty": l*w*0.15 + 2*(l+w)*d*0.15, "unit":"m3", "rate":180000},
            "Clay Brick 6 inch": {"qty": 2*(l+w)*d*60, "unit":"pc", "rate":250},
            "Cover Slab C25": {"qty": l*w*0.15, "unit":"m3", "rate":180000}
        }
    },
    "r. Soak Pit": {
        "items": ["Excavation", "Stones", "Cover"],
        "inputs": ["Diameter m", "Depth m"],
        "formula": lambda dia,d: {
            "Excavation": {"qty": 3.1416*(dia/2)**2*d, "unit":"m3", "rate":3500},
            "Stone 40mm": {"qty": 3.1416*(dia/2)**2*d*0.8, "unit":"m3", "rate":30000},
            "Concrete Cover C20": {"qty": 3.1416*(dia/2)**2*0.1, "unit":"m3", "rate":170000}
        }
    },
    "s. Boundary Wall": {
        "items": ["Foundation", "Blocks 6 inch", "Mortar"],
        "inputs": ["Length m", "Height m"],
        "formula": lambda l,h: {
            "Concrete C20": {"qty": l*0.6*0.3, "unit":"m3", "rate":170000},
            "Clay Brick 6 inch": {"qty": l*h*60*1.05, "unit":"pc", "rate":250},
            "Mortar 1:4": {"qty": l*h*0.025, "unit":"m3", "rate":95000}
        }
    },
    "t. Water Tank Base": {
        "items": ["Concrete C30", "Steel Y16", "Waterproofing"],
        "inputs": ["Length m", "Width m", "Thickness m"],
        "formula": lambda l,w,t: {
            "Concrete C30": {"qty": l*w*t, "unit":"m3", "rate":210000},
            "Y16 Steel": {"qty": l*w*t*100, "unit":"kg", "rate":1250},
            "Cementitious Waterproof": {"qty": l*w*0.02, "unit":"bag", "rate":35000}
        }
    }
}

# ============= MATERIAL DATABASE - EDITABLE =============
if 'materials' not in st.session_state:
    st.session_state['materials'] = pd.DataFrame([
        {"Section":"CEMENT","Item":"CEM II 42.5N","Unit":"bag","Rate":12500},
        {"Section":"CEMENT","Item":"CEM I 52.5N","Unit":"bag","Rate":14500},
        {"Section":"CEMENT","Item":"White Cement","Unit":"bag","Rate":22000},
        {"Section":"CEMENT","Item":"Quick Set Cement","Unit":"bag","Rate":15000},
        {"Section":"CEMENT","Item":"Portland Cement","Unit":"bag","Rate":13500},
        {"Section":"SAND","Item":"River Sand","Unit":"m3","Rate":25000},
        {"Section":"SAND","Item":"Crushed Sand","Unit":"m3","Rate":30000},
        {"Section":"SAND","Item":"Plaster Sand","Unit":"m3","Rate":28000},
        {"Section":"SAND","Item":"Filling Sand","Unit":"m3","Rate":22000},
        {"Section":"SAND","Item":"Quarry Dust","Unit":"m3","Rate":18000},
        {"Section":"STONE","Item":"Stone 20mm","Unit":"m3","Rate":32000},
        {"Section":"STONE","Item":"Stone 40mm","Unit":"m3","Rate":30000},
        {"Section":"STONE","Item":"Ballast","Unit":"m3","Rate":28000},
        {"Section":"STONE","Item":"Hardcore","Unit":"m3","Rate":20000},
        {"Section":"STONE","Item":"Aggregate Dust","Unit":"m3","Rate":25000},
        {"Section":"STONE","Item":"Chippings","Unit":"m3","Rate":35000},
        {"Section":"BRICK","Item":"Clay Brick 6 inch","Unit":"pc","Rate":250},
        {"Section":"BRICK","Item":"Clay Brick 4 inch","Unit":"pc","Rate":180},
        {"Section":"BRICK","Item":"Concrete Block 23cm Solid","Unit":"pc","Rate":1200},
        {"Section":"BRICK","Item":"Concrete Block 15cm Solid","Unit":"pc","Rate":950},
        {"Section":"BRICK","Item":"Concrete Block 23cm Hollow","Unit":"pc","Rate":850},
        {"Section":"BRICK","Item":"Concrete Block 15cm Hollow","Unit":"pc","Rate":650},
        {"Section":"BRICK","Item":"Paving Block","Unit":"m2","Rate":8500},
        {"Section":"BRICK","Item":"Face Brick","Unit":"pc","Rate":450},
        {"Section":"STEEL_BAR","Item":"R6 Stirrups","Unit":"kg","Rate":1350},
        {"Section":"STEEL_BAR","Item":"R8 Stirrups","Unit":"kg","Rate":1320},
        {"Section":"STEEL_BAR","Item":"R10 Steel","Unit":"kg","Rate":1310},
        {"Section":"STEEL_BAR","Item":"Y12 Steel","Unit":"kg","Rate":1300},
        {"Section":"STEEL_BAR","Item":"Y16 Steel","Unit":"kg","Rate":1250},
        {"Section":"STEEL_BAR","Item":"Y20 Steel","Unit":"kg","Rate":1260},
        {"Section":"STEEL_BAR","Item":"Y25 Steel","Unit":"kg","Rate":1280},
        {"Section":"STEEL_BAR","Item":"Y32 Steel","Unit":"kg","Rate":1350},
        {"Section":"STEEL_BAR","Item":"Y40 Steel","Unit":"kg","Rate":1450},
        {"Section":"STEEL_BAR","Item":"Y50 Steel","Unit":"kg","Rate":1550},
        {"Section":"BINDING","Item":"Binding Wire","Unit":"kg","Rate":2000},
        {"Section":"BINDING","Item":"BRC Mesh A142","Unit":"m2","Rate":1700},
        {"Section":"BINDING","Item":"BRC Mesh A98","Unit":"m2","Rate":1350},
        {"Section":"BINDING","Item":"Chain Link","Unit":"m","Rate":8500},
        {"Section":"BINDING","Item":"Barbed Wire","Unit":"kg","Rate":3200},
        {"Section":"TIMBER","Item":"Cypress 2x2","Unit":"m","Rate":450},
        {"Section":"TIMBER","Item":"Cypress 3x2","Unit":"m","Rate":650},
        {"Section":"TIMBER","Item":"Cypress 4x2","Unit":"m","Rate":850},
        {"Section":"TIMBER","Item":"Plywood 18mm","Unit":"pc","Rate":45000},
        {"Section":"TIMBER","Item":"Marine Board 18mm","Unit":"pc","Rate":65000},
        {"Section":"TIMBER","Item":"MDF Board","Unit":"pc","Rate":35000},
        {"Section":"TIMBER","Item":"Hardwood","Unit":"m3","Rate":450000},
        {"Section":"TIMBER","Item":"Softwood","Unit":"m3","Rate":280000},
        {"Section":"TIMBER","Item":"Pine","Unit":"m3","Rate":320000},
        {"Section":"IRON_SHEET","Item":"Mabati G30 Versatile","Unit":"m2","Rate":8500},
        {"Section":"IRON_SHEET","Item":"Mabati G28 IT4","Unit":"m2","Rate":9200},
        {"Section":"IRON_SHEET","Item":"Mabati G30 Corrugated","Unit":"m2","Rate":7800},
        {"Section":"IRON_SHEET","Item":"Resincot","Unit":"m2","Rate":12000},
        {"Section":"IRON_SHEET","Item":"Gauge 28","Unit":"m2","Rate":9500},
        {"Section":"IRON_SHEET","Item":"Gauge 30","Unit":"m2","Rate":8200},
        {"Section":"IRON_SHEET","Item":"Aluminium Sheet","Unit":"m2","Rate":15000},
        {"Section":"PAINT","Item":"Emulsion Paint","Unit":"L","Rate":8500},
        {"Section":"PAINT","Item":"Weather Guard","Unit":"L","Rate":12000},
        {"Section":"PAINT","Item":"Undercoat","Unit":"L","Rate":7500},
        {"Section":"PAINT","Item":"Varnish","Unit":"L","Rate":15000},
        {"Section":"PAINT","Item":"Oil Based","Unit":"L","Rate":9500},
        {"Section":"PAINT","Item":"Water Based","Unit":"L","Rate":8500},
        {"Section":"PAINT","Item":"Primer","Unit":"L","Rate":7000},
        {"Section":"TILES","Item":"Ceramic Tile 30x30","Unit":"m2","Rate":15000},
        {"Section":"TILES","Item":"Ceramic Tile 60x60","Unit":"m2","Rate":22000},
        {"Section":"TILES","Item":"Porcelain Tile","Unit":"m2","Rate":28000},
        {"Section":"TILES","Item":"Granite Tile","Unit":"m2","Rate":45000},
        {"Section":"TILES","Item":"Marble Tile","Unit":"m2","Rate":65000},
        {"Section":"TILES","Item":"PVC Tile","Unit":"m2","Rate":12000},
        {"Section":"TILES","Item":"Terrazzo","Unit":"m2","Rate":35000},
        {"Section":"CONCRETE","Item":"Concrete C15","Unit":"m3","Rate":160000},
        {"Section":"CONCRETE","Item":"Concrete C20","Unit":"m3","Rate":170000},
        {"Section":"CONCRETE","Item":"Concrete C25","Unit":"m3","Rate":180000},
        {"Section":"CONCRETE","Item":"Concrete C30","Unit":"m3","Rate":210000},
        {"Section":"CONCRETE","Item":"Concrete C35","Unit":"m3","Rate":230000},
        {"Section":"CONCRETE","Item":"Concrete C40","Unit":"m3","Rate":250000},
        {"Section":"CONCRETE","Item":"Ready Mix C25","Unit":"m3","Rate":195000},
        {"Section":"CONCRETE","Item":"Ready Mix C30","Unit":"m3","Rate":225000},
        {"Section":"GLASS","Item":"Glass 4mm Clear","Unit":"m2","Rate":8500},
        {"Section":"GLASS","Item":"Glass 5mm Tinted","Unit":"m2","Rate":12000},
        {"Section":"GLASS","Item":"Glass 6mm Laminated","Unit":"m2","Rate":18000},
        {"Section":"GLASS","Item":"Mirror","Unit":"m2","Rate":15000},
        {"Section":"GLASS","Item":"Frosted Glass","Unit":"m2","Rate":14000},
        {"Section":"GLASS","Item":"Bullet Proof","Unit":"m2","Rate":250000},
        {"Section":"PIPES","Item":"PPR 20mm","Unit":"m","Rate":2500},
        {"Section":"PIPES","Item":"PPR 25mm","Unit":"m","Rate":3500},
        {"Section":"PIPES","Item":"PPR 32mm","Unit":"m","Rate":4800},
        {"Section":"PIPES","Item":"PPR 40mm","Unit":"m","Rate":6500},
        {"Section":"PIPES","Item":"PVC 4inch Waste","Unit":"m","Rate":4500},
        {"Section":"PIPES","Item":"PVC 2inch Waste","Unit":"m","Rate":2800},
        {"Section":"PIPES","Item":"PVC 6inch","Unit":"m","Rate":8500},
        {"Section":"PIPES","Item":"HDPE 50mm","Unit":"m","Rate":5500},
        {"Section":"PIPES","Item":"GI Pipe 1inch","Unit":"m","Rate":3800},
        {"Section":"WIRES","Item":"Cable 1.5mm","Unit":"m","Rate":850},
        {"Section":"WIRES","Item":"Cable 2.5mm","Unit":"m","Rate":1200},
        {"Section":"WIRES","Item":"Cable 4mm","Unit":"m","Rate":1800},
        {"Section":"WIRES","Item":"Cable 6mm","Unit":"m","Rate":2500},
        {"Section":"WIRES","Item":"Cable 10mm","Unit":"m","Rate":4200},
        {"Section":"WIRES","Item":"Cable 16mm","Unit":"m","Rate":6500},
        {"Section":"WIRES","Item":"Twin Earth","Unit":"m","Rate":2200},
        {"Section":"WIRES","Item":"Armoured Cable","Unit":"m","Rate":12000},
        {"Section":"DOORS","Item":"Steel Door 90x210","Unit":"pc","Rate":95000},
        {"Section":"DOORS","Item":"Flush Door","Unit":"pc","Rate":45000},
        {"Section":"DOORS","Item":"Panel Door","Unit":"pc","Rate":65000},
        {"Section":"DOORS","Item":"PVC Door","Unit":"pc","Rate":55000},
        {"Section":"DOORS","Item":"Glass Door","Unit":"pc","Rate":120000},
        {"Section":"DOORS","Item":"Security Door","Unit":"pc","Rate":250000},
        {"Section":"WINDOWS","Item":"Aluminium Sliding","Unit":"m2","Rate":65000},
        {"Section":"WINDOWS","Item":"Aluminium Casement","Unit":"m2","Rate":72000},
        {"Section":"WINDOWS","Item":"Louver Window","Unit":"m2","Rate":35000},
        {"Section":"WINDOWS","Item":"Wooden Window","Unit":"m2","Rate":55000},
        {"Section":"WINDOWS","Item":"UPVC Window","Unit":"m2","Rate":85000},
        {"Section":"WATERPROOF","Item":"APP Membrane","Unit":"m2","Rate":8500},
        {"Section":"WATERPROOF","Item":"Bitumen Paint","Unit":"L","Rate":4500},
        {"Section":"WATERPROOF","Item":"Cementitious Waterproof","Unit":"bag","Rate":35000},
        {"Section":"WATERPROOF","Item":"Liquid Membrane","Unit":"L","Rate":12000},
        {"Section":"WATERPROOF","Item":"EPDM","Unit":"m2","Rate":15000},
        {"Section":"CEILING","Item":"Gypsum Board","Unit":"pc","Rate":18000},
        {"Section":"CEILING","Item":"PVC Ceiling","Unit":"m2","Rate":12000},
        {"Section":"CEILING","Item":"Acoustic Ceiling","Unit":"m2","Rate":25000},
        {"Section":"CEILING","Item":"Wooden T&G","Unit":"m2","Rate":18000},
        {"Section":"CEILING","Item":"Suspended Ceiling","Unit":"m2","Rate":22000},
        {"Section":"CEILING","Item":"Mineral Fiber","Unit":"m2","Rate":28000},
        {"Section":"NAILS","Item":"Nails 1 inch","Unit":"kg","Rate":2200},
        {"Section":"NAILS","Item":"Nails 2 inch","Unit":"kg","Rate":2200},
        {"Section":"NAILS","Item":"Nails 3 inch","Unit":"kg","Rate":2200},
        {"Section":"NAILS","Item":"Nails 4 inch","Unit":"kg","Rate":2200},
        {"Section":"NAILS","Item":"Roofing Nails","Unit":"kg","Rate":2500},
        {"Section":"NAILS","Item":"Concrete Nails","Unit":"kg","Rate":2800},
        {"Section":"NAILS","Item":"U-Nails","Unit":"kg","Rate":3200},
        {"Section":"FITTINGS","Item":"WC Complete","Unit":"pc","Rate":180000},
        {"Section":"FITTINGS","Item":"Wash Basin","Unit":"pc","Rate":95000},
        {"Section":"FITTINGS","Item":"Kitchen Sink","Unit":"pc","Rate":75000},
        {"Section":"FITTINGS","Item":"Shower Mixer","Unit":"pc","Rate":65000},
        {"Section":"FITTINGS","Item":"Towel Rail","Unit":"pc","Rate":25000},
        {"Section":"FITTINGS","Item":"PPR Fittings","Unit":"pc","Rate":1500},
        {"Section":"ELECTRICAL","Item":"Socket 13A","Unit":"pc","Rate":4500},
        {"Section":"ELECTRICAL","Item":"Light Point","Unit":"pc","Rate":8500},
        {"Section":"ELECTRICAL","Item":"Switch","Unit":"pc","Rate":3500},
        {"Section":"ELECTRICAL","Item":"DB Board","Unit":"pc","Rate":125000},
        {"Section":"ELECTRICAL","Item":"Earth Rod","Unit":"pc","Rate":45000},
        {"Section":"DRAINAGE","Item":"Manhole","Unit":"pc","Rate":125000},
        {"Section":"EXTERNAL","Item":"Road Tarmac","Unit":"m2","Rate":35000},
        {"Section":"EXTERNAL","Item":"Chain Link Fence","Unit":"m","Rate":8500},
        {"Section":"EXTERNAL","Item":"Grass","Unit":"m2","Rate":2500},
        {"Section":"FORMWORK","Item":"Formwork","Unit":"m2","Rate":8500},
        {"Section":"FINISHING","Item":"Plaster 1:4","Unit":"m2","Rate":4500},
        {"Section":"FINISHING","Item":"Screed 1:4","Unit":"m2","Rate":8500},
        {"Section":"FINISHING","Item":"Terrazzo","Unit":"m2","Rate":35000},
        {"Section":"FINISHING","Item":"Mortar 1:4","Unit":"m3","Rate":95000},
    ])

# ============= SESSION STATE =============
if 'measurements' not in st.session_state:
    st.session_state['measurements'] = []
if 'drawing_image' not in st.session_state:
    st.session_state['drawing_image'] = None
if 'drawing_pdf' not in st.session_state:
    st.session_state['drawing_pdf'] = None
if 'takeoff_list' not in st.session_state:
    st.session_state['takeoff_list'] = []
if 'selected_main' not in st.session_state:
    st.session_state['selected_main'] = "CEMENT"
if 'selected_sub' not in st.session_state:
    st.session_state['selected_sub'] = None

tabs = st.tabs(["1. 📐 Drawing Upload", "2. 🏗️ 20 Elements Takeoff", "3. 🧱 Material Browser", "4. 📊 BOQ + Excel Export"])

# ===== TAB 1: DRAWING UPLOAD - PDF FIXED =====
with tabs[0]:
    st.header("📐 Drawing Upload - For Excel Reference")
    uploaded_file = st.file_uploader("Upload PDF/Image ya Plan", type=['png','jpg','jpeg','pdf'])
    if uploaded_file:
        if uploaded_file.type == "application/pdf":
            st.info("📄 PDF uploaded: " + uploaded_file.name)
            st.session_state['drawing_pdf'] = uploaded_file
            st.success("✅ PDF yashyizweho. Izashyirwa muri Excel!")
        else:
            image = Image.open(uploaded_file)
            st.session_state['drawing_image'] = image
            st.image(image, caption="Drawing for Excel Export", use_column_width=True)
            st.success("✅ Drawing yashyizweho. Izajya muri Excel!")

    st.divider()
    st.subheader("Manual Takeoff List")
    col1, col2, col3, col4 = st.columns(4)
    with col1: desc = st.text_input("Description", "Footing F1", key="takeoff_desc")
    with col2: qty = st.number_input("Quantity", min_value=0.0, value=0.0, step=0.1, key="takeoff_qty")
    with col3: unit = st.selectbox("Unit", ["m","m2","m3","pc","kg","L"], key="takeoff_unit")
    with col4:
        st.write("")
        if st.button("Add to List", key="takeoff_add"):
            st.session_state['takeoff_list'].append({"Description":desc, "Qty":qty, "Unit":unit})
            st.success("Added!")

    if st.session_state['takeoff_list']:
        st.dataframe(pd.DataFrame(st.session_state['takeoff_list']), hide_index=True, use_container_width=True)  # ===== TAB 2: 20 ELEMENTS TAKEOFF =====
with tabs[1]:
    st.header("🏗️ 20 Elements Takeoff - SMM Standard")
    st.caption("Hitamo Element → Andika Dimensions → Ihita ibara Materials zose + Cost")

    element_name = st.selectbox("Hitamo Element (20 Total)", list(ELEMENTS.keys()), key="element_select")
    element_data = ELEMENTS[element_name]

    st.subheader(f"{element_name}")
    st.write(f"*Materials:* {', '.join(element_data['items'])}")

    inputs = {}
    cols = st.columns(len(element_data['inputs']))
    for i, inp in enumerate(element_data['inputs']):
        with cols[i]:
            if "Type" in inp:
                inputs[inp] = st.selectbox(inp, ["6 inch", "4 inch"], key=f"input_{i}_{element_name}")
            elif "Dia" in inp:
                inputs[inp] = st.selectbox(inp, [16, 20, 25], key=f"input_{i}_{element_name}")
            elif "Number" in inp or "pcs" in inp:
                inputs[inp] = st.number_input(inp, min_value=1, value=1, step=1, key=f"input_{i}_{element_name}")
            else:
                inputs[inp] = st.number_input(inp, min_value=0.1, value=1.0, step=0.1, key=f"input_{i}_{element_name}")

    if st.button("Calculate Element", type="primary", key=f"calc_{element_name}"):
        calc_values = list(inputs.values())
        results = element_data['formula'](*calc_values)

        st.subheader("Results:")
        total_cost = 0
        for item, data in results.items():
            if data['qty'] > 0:
                amount = data['qty'] * data['rate']
                total_cost += amount
                st.session_state['measurements'].append({
                    "Element": element_name.split(".")[1].strip(),
                    "Item": item,
                    "Qty": round(data['qty'],2),
                    "Unit": data['unit'],
                    "Rate": data['rate'],
                    "Amount": round(amount,0),
                    "Source": "Element Takeoff"
                })
                st.write(f"✅ *{item}: {data['qty']:.2f} {data['unit']} x {data['rate']:,} = *{amount:,.0f} RWF**")

        st.success(f"*TOTAL for {element_name}: {total_cost:,.0f} RWF* - Added to BOQ!")
        st.balloons()

# ===== TAB 3: MATERIAL BROWSER - EDITABLE RATES =====
with tabs[2]:
    st.header("🧱 Material Browser - Editable Rates")
    st.info("Hitamo Main Material → Kanda kuri Rate uyihindure → Andika Qty → Add to BOQ")

    col1, col2 = st.columns([1,3])

    with col1:
        st.subheader("Main Materials")
        main_material = st.radio("Hitamo Material:", list(MATERIAL_TREE.keys()), key="main_radio")
        st.session_state['selected_main'] = main_material

    with col2:
        st.subheader(f"Table ya {st.session_state['selected_main']} - Kanda kuri Rate uyihindure")
        if st.session_state['selected_main']:
            # Filter materials for selected section
            mat_df = st.session_state['materials'][st.session_state['materials']['Section']==st.session_state['selected_main']].copy()
            mat_df = mat_df[['Item','Unit','Rate']].reset_index(drop=True)
            
            # EDITABLE DATAFRAME - Rate ishobora guhinduka
            edited_df = st.data_editor(
                mat_df,
                column_config={
                    "Item": st.column_config.TextColumn("Item", disabled=True),
                    "Unit": st.column_config.TextColumn("Unit", disabled=True),
                    "Rate": st.column_config.NumberColumn("Rate RWF", min_value=0, step=100, format="%d")
                },
                hide_index=True,
                use_container_width=True,
                key="editable_rates"
            )
            
            # Update session state with edited rates
            for idx, row in edited_df.iterrows():
                mask = (st.session_state['materials']['Section']==st.session_state['selected_main']) & (st.session_state['materials']['Item']==row['Item'])
                st.session_state['materials'].loc[mask, 'Rate'] = row['Rate']
            
            st.divider()
            st.subheader("Hitamo Material Wongere kuri BOQ")
            
            selected_item = st.selectbox("Hitamo Item:", edited_df['Item'].tolist(), key="item_select")
            
            if selected_item:
                mat = st.session_state['materials'][
                    (st.session_state['materials']['Section']==st.session_state['selected_main']) &
                    (st.session_state['materials']['Item']==selected_item)
                ].iloc[0]
                
                col1, col2, col3 = st.columns(3)
                col1.metric("Unit", mat['Unit'])
                col2.metric("Rate", f"{mat['Rate']:,} RWF")
                col3.metric("Status", "✅ Editable")

                qty = st.number_input("Andika Quantity", min_value=0.0, value=0.0, max_value=999999999.0, step=0.1, key="qty_input")

                if st.button("➕ Add to BOQ", type="primary", use_container_width=True):
                    st.session_state['measurements'].append({
                        "Element": "Manual Selection",
                        "Item": selected_item,
                        "Qty": qty,
                        "Unit": mat['Unit'],
                        "Rate": mat['Rate'],
                        "Amount": round(qty * mat['Rate'],0),
                        "Source": "Material Browser"
                    })
                    st.success(f"Added {qty} {mat['Unit']} of {selected_item} @ {mat['Rate']:,} = {qty*mat['Rate']:,.0f} RWF")
                    st.balloons()

# ===== TAB 4: BOQ + EXCEL EXPORT =====
with tabs[3]:
    st.header("📊 Bill of Quantities + Excel Export")

    if not st.session_state['measurements']:
        st.warning("Nta materials zirimo. Koresha Tab 2 cyangwa Tab 3 wongere Elements/Materials.")
    else:
        df_boq = pd.DataFrame(st.session_state['measurements'])
        df_boq['Amount'] = df_boq['Qty'] * df_boq['Rate']

        st.subheader("BOQ Summary by Element")
        element_summary = df_boq.groupby('Element')['Amount'].sum().reset_index().sort_values('Amount', ascending=False)
        st.dataframe(element_summary, hide_index=True, use_container_width=True)

        st.divider()
        st.subheader("Detailed BOQ")
        st.dataframe(df_boq[['Element','Item','Qty','Unit','Rate','Amount','Source']], hide_index=True, use_container_width=True)

        subtotal = df_boq['Amount'].sum()
        vat = subtotal * 0.18
        total = subtotal + vat

        c1, c2, c3, c4 = st.columns(4)
        c1.metric("Subtotal", f"{subtotal:,.0f} RWF")
        c2.metric("VAT 18%", f"{vat:,.0f} RWF")
        c3.metric("GRAND TOTAL", f"{total:,.0f} RWF")
        c4.metric("Elements", f"{len(df_boq['Element'].unique())}")

        col1, col2 = st.columns(2)
        with col1:
            fig1 = px.pie(df_boq, values='Amount', names='Element', title="Cost Distribution by Element")
            st.plotly_chart(fig1, use_container_width=True)
        with col2:
            fig2 = px.bar(element_summary, x='Element', y='Amount', title="Cost by Element")
            st.plotly_chart(fig2, use_container_width=True)

        st.divider()

        def create_excel():
            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df_boq.to_excel(writer, index=False, sheet_name='BOQ', startrow=2)
                ws_boq = writer.sheets['BOQ']
                ws_boq['A1'] = "B-ESTAMER 5.7 - BILL OF QUANTITIES - 20 ELEMENTS"
                ws_boq['A1'].font = Font(bold=True, size=16, color="FFFFFF")
                ws_boq['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                ws_boq.merge_cells('A1:H1')
                ws_boq['A1'].alignment = Alignment(horizontal='center')

                if st.session_state['takeoff_list']:
                    pd.DataFrame(st.session_state['takeoff_list']).to_excel(writer, index=False, sheet_name='Takeoff', startrow=2)
                    ws_takeoff = writer.sheets['Takeoff']
                    ws_takeoff['A1'] = "MANUAL TAKEOFF LIST"
                    ws_takeoff['A1'].font = Font(bold=True, size=14)

                if st.session_state['drawing_image']:
                    ws_draw = writer.book.create_sheet('Drawings')
                    ws_draw['A1'] = "PROJECT DRAWINGS"
                    ws_draw['A1'].font = Font(bold=True, size=16, color="FFFFFF")
                    ws_draw['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    ws_draw.merge_cells('A1:D1')

                    img_byte_arr = BytesIO()
                    st.session_state['drawing_image'].save(img_byte_arr, format='PNG')
                    img_byte_arr.seek(0)
                    img = XLImage(img_byte_arr)
                    img.width = 700
                    img.height = 500
                    ws_draw.add_image(img, 'A3')
                    ws_draw['A2'] = f"Drawing uploaded: {pd.Timestamp.now().strftime('%Y-%m-%d %H:%M')}"
                
                if st.session_state['drawing_pdf']:
                    ws_pdf = writer.book.create_sheet('PDF_Reference')
                    ws_pdf['A1'] = "PDF DRAWING REFERENCE"
                    ws_pdf['A1'].font = Font(bold=True, size=16)
                    ws_pdf['A2'] = f"PDF File: {st.session_state['drawing_pdf'].name}"
                    ws_pdf['A3'] = "Note: PDF attached to this Excel file"

                ws_sum = writer.book.create_sheet('Summary')
                ws_sum['A1'] = "PROJECT COST SUMMARY"
                ws_sum['A1'].font = Font(bold=True, size=16, color="FFFFFF")
                ws_sum['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                ws_sum.merge_cells('A1:C1')

                ws_sum['A3'] = "Subtotal:"
                ws_sum['B3'] = subtotal
                ws_sum['B3'].number_format = '#,##0'
                ws_sum['A4'] = "VAT 18%:"
                ws_sum['B4'] = vat
                ws_sum['B4'].number_format = '#,##0'
                ws_sum['A5'] = "GRAND TOTAL:"
                ws_sum['B5'] = total
                ws_sum['B5'].font = Font(bold=True, size=14)
                ws_sum['B5'].number_format = '#,##0'

                ws_sum['A7'] = "Total Elements Used:"
                ws_sum['B7'] = len(df_boq['Element'].unique())
                ws_sum['A8'] = "Total Line Items:"
                ws_sum['B8'] = len(df_boq)
                ws_sum['A9'] = "Generated:"
                ws_sum['B9'] = pd.Timestamp.now().strftime('%Y-%m-%d %H:%M')

            return output.getvalue()

        st.download_button(
            "📊 Download Complete Excel (BOQ + Drawing + Takeoff)",
            create_excel(),
            "B-ESTAMER_5.7_20_ELEMENTS_BOQ.xlsx",
            mime="application/vnd.ms-excel",
            type="primary",
            use_container_width=True
        )

st.divider()
st.caption("B-ESTAMER 5.7 | SMM Standard | 0787993679")

